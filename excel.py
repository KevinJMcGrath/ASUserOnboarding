from openpyxl import load_workbook

xls_to_sfdc = {
        "Contact ID": "Id",
        "Account ID": "AccountId",
        "First Name": "FirstName",
        "Last Name": "LastName",
        "Email": "Email",
        "Type": "Type__c",
        "Username_AS": "Username__c",
        "Password_AS": "Password__c",
        "Confirm Password_AS": "Confirm_Password__c",
        "Effective Date (paying user)": "Effective_Date_paying_user__c",
        "Current Contract Effective Date": "Current_Contract_Effective_Date__c",
        "User Status (contract)": "User_Status__c",
        "User Status (platform)_AS": "Status_of_User__c",
        "Primary Research (PRM)_AS": "Region_s_Subscribed__c",
        "Transcripts Type_AS": "Transcripts_Type__c",
        "LexisNexis_AS": "LexisNexis__c",
        "AS Premium News Content": "AS_Premium_News_Content__c",
        "BRM User Type": "TR_Request_Type__c",
        "BRM Purchase Date": "TR_Service_Start_Date__c",
        "Broker Research (BRM) (TR) Region_AS": "Broker_Research_Module__c",
        "Opportunity": "Opportunity__c",
        "Renewal Opportunity": "Renewal_Opportunity__c",
        "Dashboard Access_AS": "Feature_Flag_2_AS__c",
        "Dashboard Auto Alerts_AS": "Dashboard_Auto_Alerts_AS__c",
        "MTNewswireLiveBriefPro_AS": "MTNewswireLiveBriefPro_AS__c",
        "PositionId_AS": "PositionId__c",
        "Trial Start_AS": "Trial_Start__c",
        "Trial Kick-Off": "Trial_Kick_Off__c",
        "Trial End_AS": "Trial_End__c",
        "Contact Owner": "OwnerId",
        "Mailing Street": "MailingStreet",
        "Mailing City": "MailingCity",
        "Mailing State/Province": "MailingState",
        "Mailing Zip/Postal Code": "MailingPostalCode",
        "Mailing Country": "MailingCountry",
        "Lead Source": "LeadSource"
    }

def load_workbook_data(wb_path):
    template_ws_name = 'Upload Template'
    user_ws_name = 'New Users'
    opp_ws_name = 'Opportunity Update'

    workbook = load_workbook(filename=wb_path, read_only=True, data_only=True)

    ws_user_sheet = workbook[user_ws_name]
    user_dict = []
    for row in ws_user_sheet.iter_rows(min_row=2, values_only=True):
        user_email = row[2]
        u = {
            "firstname": row[0],
            "lastname": row[1],
            "email": user_email
        }

        user_dict[user_email] = u


    ws_opp_sheet = workbook[opp_ws_name]
    opp_id = ''
    opp_role = ''

    for row in ws_opp_sheet.iter_rows(min_row=2, max_row=3, values_only=True):
        opp_id = row[0]
        opp_role = row[2]
        break


    ws_upload_template = workbook[template_ws_name]

    row_list = []
    for row in ws_upload_template.iter_rows(min_row=1, max_row=2, values_only=True):
        row_list.append(list(row))

    template_values = dict(zip(row_list[0], row_list[1]))

    return template_values, user_dict, opp_id, opp_role
